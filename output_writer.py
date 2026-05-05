"""
כותבי פלט — מייצרים את שני קבצי ה-Excel הממולאים עם קידוד צבעים.

positions_filled.xlsx
  • אותה פריסה כמו קובץ העמדות הקלט.
  • תאי "1" פתוחים מוחלפים בשם העובד המוקצה.
  • תאים נעולים (מוקצים מראש) אינם נגעים.

forecast_filled.xlsx
  • אותה פריסה כמו קובץ הצפי הקלט.
  • תאים שבהם העובד הוקצה מציגים את שם העמדה.
    - אם התא כבר היה לא-ריק (למשל "AL"), העמדה מצורפת.
    - אם ההקצאה הפרה אילוץ חסום (X) → תא אדום.
  • עובדים ללא רשומת אילוצים → כל השורה ברקע צהוב.
"""
from __future__ import annotations

from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from models import DAYS, SHIFT_ORDER, Employee, PositionSlot, ShiftType


# ---------------------------------------------------------------------------
# קבועי צבע
# ---------------------------------------------------------------------------

RED_FILL    = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFB347", end_color="FFB347", fill_type="solid")
GRAY_FILL   = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
RED_FONT    = Font(color="CC0000", bold=True)
BOLD_FONT   = Font(bold=True)


# ---------------------------------------------------------------------------
# פלט עמדות
# ---------------------------------------------------------------------------

def write_positions_filled(
    slots: List[PositionSlot],
    source_path: str,
    output_path: str,
    *,
    pos_col: int = 1,        # 1-based: עמודה A = שם עמדה
    shift_col: int = 3,      # 1-based: עמודה C = סוג משמרת (אחרי Post + Type)
    day_start_col: int = 4,  # 1-based: עמודה D = יום ראשון (D=Sun … J=Sat)
    header_row: int = 2,     # שורה 2 היא הכותרת (שורה 1 ריקה)
) -> None:
    """מחליף סלוטים פתוחים בקובץ העמדות בשמות העובדים המוקצים.

    עבור תאים שהערך המקורי שלהם היה מספר > 1 (למשל RELIEF GUARD עם "2"
    או "3"), שורות נוספות מוכנסות מתחת לשורה המקורית כדי שכל עובד מוקצה
    יופיע בשורה משלו, בהתאמה לפריסה של עמדות אחרות.
    """
    from collections import defaultdict
    from copy import copy as _copy

    wb = load_workbook(source_path)
    ws = wb.active

    # בניית מפת בדיקה: (excel_row, day) -> [emp1, emp2, ...]  (מסודר, סלוטים פתוחים בלבד)
    multi_map: Dict[Tuple[int, int], List[str]] = defaultdict(list)
    ot_set: set = set()
    for slot in slots:
        if slot.locked or not slot.assigned_employee or slot.excel_row == 0:
            continue
        multi_map[(slot.excel_row, slot.day)].append(slot.assigned_employee)
        if slot.is_ot:
            ot_set.add((slot.excel_row, slot.day, slot.assigned_employee))

    # --- שלב 1: כתיבת העובד הראשון לכל תא קיים ---
    for (excel_row, day), emps in multi_map.items():
        col  = day_start_col + day
        cell = ws.cell(row=excel_row, column=col)
        emp_name = emps[0]
        is_ot = (excel_row, day, emp_name) in ot_set
        cell.value = f"{emp_name} OT" if is_ot else emp_name
        cell.font  = BOLD_FONT
        if is_ot:
            cell.fill = ORANGE_FILL

    # --- שלב 2: איתור שורות שדורשות הרחבה (יותר מעובד אחד ליום) ---
    # מקסימום עובדים מוקצים לכל יום עבור כל excel_row
    row_max: Dict[int, int] = {}
    for (excel_row, day), emps in multi_map.items():
        row_max[excel_row] = max(row_max.get(excel_row, 0), len(emps))

    # עיבוד בסדר הפוך של שורות כדי שהכנסות מוקדמות לא יזיזו אינדקסים
    for excel_row in sorted(row_max.keys(), reverse=True):
        n_emps = row_max[excel_row]
        if n_emps <= 1:
            continue

        n_extra = n_emps - 1
        insert_at = excel_row + 1
        ws.insert_rows(insert_at, amount=n_extra)

        # העתקת עיצוב מהשורה המקורית לכל שורה חדשה
        orig_cells = list(ws[excel_row])
        for extra_i in range(n_extra):
            new_row_idx = insert_at + extra_i
            for orig_cell in orig_cells:
                new_cell = ws.cell(row=new_row_idx, column=orig_cell.column)
                # העתקת עמודות סטטיות (שם עמדה, סוג, משמרת)
                if orig_cell.column < day_start_col:
                    new_cell.value = orig_cell.value
                # העתקת גבולות כדי שהשורה החדשה תשתלב
                if orig_cell.border:
                    new_cell.border = _copy(orig_cell.border)

        # מילוי עובדים נוספים בשורות החדשות
        for (er, day), emps in multi_map.items():
            if er != excel_row:
                continue
            for extra_i, emp_name in enumerate(emps[1:]):
                new_row_idx = insert_at + extra_i
                col  = day_start_col + day
                cell = ws.cell(row=new_row_idx, column=col)
                is_ot = (er, day, emp_name) in ot_set
                cell.value = f"{emp_name} OT" if is_ot else emp_name
                cell.font  = BOLD_FONT
                if is_ot:
                    cell.fill = ORANGE_FILL

    wb.save(output_path)
    print(f"[Output] Saved: {output_path}")


# ---------------------------------------------------------------------------
# פלט צפי
# ---------------------------------------------------------------------------

def write_forecast_filled(
    employees: List[Employee],
    source_path: str,
    output_path: str,
    *,
    name_col: int = 1,       # 1-based: עמודה A
    shift_start_col: int = 3,  # 1-based: עמודת המשמרת הראשונה (עמודה C, אחרי שם+ותק)
    header_row: int = 1,
) -> None:
    """מכסה הקצאות על קובץ הצפי עם קידוד צבעים."""
    wb = load_workbook(source_path)
    ws = wb.active

    # בניית מפת בדיקה של עובדים לפי שם באותיות קטנות
    emp_map = {e.name.strip().lower(): e for e in employees}

    for row in range(header_row + 1, ws.max_row + 1):
        name_cell = ws.cell(row=row, column=name_col)
        name = str(name_cell.value or "").strip()
        if not name:
            continue

        emp = emp_map.get(name.lower())
        if emp is None:
            continue

        # צביעת השורה כולה לפי סטטוס האילוצים:
        #   ללא אילוצים + required_shifts == 0  → אדום  (לא הגיש, לא מתוזמן)
        #   ללא אילוצים + required_shifts  > 0  → צהוב (לא הגיש, אך יש לו משמרות)
        if not emp.has_constraints:
            row_fill = RED_FILL if emp.required_shifts == 0 else YELLOW_FILL
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = row_fill

        # כתיבת הקצאות, סימון הפרות, וסימון תאים לא-זמינים ב-X
        for idx in range(21):
            day   = idx // 3
            shift = SHIFT_ORDER[idx % 3]
            col   = shift_start_col + idx

            cell     = ws.cell(row=row, column=col)
            orig_val = str(cell.value or "").strip()
            if orig_val.upper() == "X":
                orig_val = ""  # X בשלב 1 הוא סמן זמינות, לא תווית תצוגה

            assigned_pos: Optional[str] = next(
                (pos for d, s, pos in emp.assigned if d == day and s == shift),
                None,
            )
            is_violation   = (day, shift) in emp.violations
            is_unavailable = (
                emp.has_constraints
                and not emp.availability.get((day, shift), False)
            )

            is_ot = (day, shift) in {(d, s) for d, s, _ in emp.ot_assignments}

            if assigned_pos is not None:
                # הוספת X לפני ההקצאה כאשר היא הפרה אילוץ חסום
                label   = f"X {assigned_pos}" if is_violation else assigned_pos
                new_val = f"{orig_val} → {label}" if orig_val else label
                cell.value = new_val
                if is_violation:
                    cell.fill = RED_FILL
                    cell.font = RED_FONT
                elif is_ot:
                    cell.fill = ORANGE_FILL
                    cell.font = BOLD_FONT
                else:
                    cell.font = BOLD_FONT
            elif is_unavailable and not orig_val:
                # תא ריק שבו העובד חסום → סימון למנהל
                cell.value = "X"

    # הוספת עמודת "משמרות בפועל" אחרי העמודה האחרונה
    actual_col = ws.max_column + 1
    header_cell = ws.cell(row=header_row, column=actual_col)
    header_cell.value = "משמרות בפועל"
    header_cell.font  = BOLD_FONT

    for row in range(header_row + 1, ws.max_row + 1):
        name_cell = ws.cell(row=row, column=name_col)
        name = str(name_cell.value or "").strip()
        if not name:
            continue
        emp = emp_map.get(name.lower())
        if emp is None:
            continue
        cell = ws.cell(row=row, column=actual_col)
        cell.value = emp.assigned_count
        if emp.assigned_count > emp.max_shifts:
            cell.fill = RED_FILL
            cell.font = RED_FONT
        elif emp.assigned_count > emp.required_shifts:
            cell.fill = ORANGE_FILL
            cell.font = BOLD_FONT

    wb.save(output_path)
    print(f"[Output] Saved: {output_path}")


# ---------------------------------------------------------------------------
# שלב 1: כתיבת צפי עם איקסים
# ---------------------------------------------------------------------------

def write_forecast_with_constraints(
    employees: List[Employee],
    forecast_path: str,
    output_path: str,
    *,
    name_col: int = 1,
    shift_start_col: int = 3,
    header_row: int = 1,
) -> None:
    """שלב 1: סימון X אפור בתאי הצפי שבהם העובד חסום לפי קובץ האילוצים.

    תאים ריקים שהעובד חסום בהם → X עם רקע אפור.
    תאים שכבר מכילים ערך (TR, AL, X ידני וכו') → לא נגעים.
    עובדים ללא רשומה בקובץ האילוצים → לא מסומנים (יהיו צהובים בשלב 2).
    """
    wb = load_workbook(forecast_path)
    ws = wb.active

    emp_map = {e.name.strip().lower(): e for e in employees}

    for row in range(header_row + 1, ws.max_row + 1):
        name_cell = ws.cell(row=row, column=name_col)
        name = str(name_cell.value or "").strip()
        if not name:
            continue

        emp = emp_map.get(name.lower())
        if emp is None or not emp.has_constraints:
            continue

        for idx in range(21):
            day   = idx // 3
            shift = SHIFT_ORDER[idx % 3]
            col   = shift_start_col + idx

            cell = ws.cell(row=row, column=col)
            if cell.value is not None and str(cell.value).strip():
                continue  # תא שכבר מכיל ערך — לא מגעים

            if not emp.availability.get((day, shift), True):
                cell.value = "X"
                cell.fill  = GRAY_FILL

    wb.save(output_path)
    print(f"[Output] Saved: {output_path}")


# ---------------------------------------------------------------------------
# פונקציית עזר פנימית (משקפת data_loader._parse_shift_type למניעת ייבוא מעגלי)
# ---------------------------------------------------------------------------

def _parse_shift_type(raw: str) -> Optional[ShiftType]:
    v = raw.lower().strip()
    if v in ("morning", "בוקר", "morn", "am", "1"):
        return ShiftType.MORNING
    if v in ("noon", "afternoon", "צהריים", "pm", "mid", "2"):
        return ShiftType.NOON
    if v in ("night", "evening", "לילה", "pm2", "3"):
        return ShiftType.NIGHT
    return None
